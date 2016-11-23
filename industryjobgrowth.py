import csv
import sqlite3
import matplotlib.pyplot as plt
import openpyxl

# Create in-memory database and initialize tables
con = sqlite3.connect(":memory:")
cur = con.cursor()
cur.execute("CREATE TABLE projections (SOC, Change, Industry);") # use your column names here
cur.execute("CREATE TABLE codelookup (Code, Occupation);") # use your column names here
cur.execute("CREATE TABLE popproj (agegrpcode INT, agegrp, yr_2015 INT, yr_2025 INT);") # use your column names here


with open('//Users/bill/School/MSIT/Data/occupationcodes.csv') as OccupationCodesFile:
    codelookup = csv.DictReader(OccupationCodesFile)
    to_db = [(i['Code'], i['Occupation']) for i in codelookup]

cur.executemany("INSERT INTO codelookup (Code, Occupation) VALUES (?, ?);", to_db)
con.commit()

cur.executescript("UPDATE codelookup SET Occupation = SUBSTR(Occupation, 1,LENGTH(Occupation)-12)")
con.commit


with open('/Users/bill/School/MSIT/Data/Long_Term_Occupational_Projections.csv') as ProjectionsFile:
    projections = csv.DictReader(ProjectionsFile)
    to_db = [(i['SOC'], i['Change']) for i in projections]

cur.executemany("INSERT INTO projections (SOC, Change) VALUES (?, ?);", to_db)
con.commit()

for row in cur.execute('Select SOC from projections'):
    cur.executescript("UPDATE projections set Industry = SUBSTR(SOC,3,2)")
    con.commit

cur.executescript("DELETE from projections WHERE Industry = '00'")
con.commit


IndustryAndTotalChange = []

for row in cur.execute('SELECT codelookup.Occupation as Code, SUM(projections.Change) as Total FROM projections '
                      'INNER JOIN codelookup ON codelookup.Code=projections.Industry GROUP BY codelookup.Occupation '
                      'ORDER BY SUM(projections.Change) DESC'):
    IndustryAndTotalChange.append(row)

for row in cur.execute('SELECT SUM(Change) FROM projections'):
    TenYearTotalGrowth = int(row[0])


print("Ten Year Projected Job Growth")
print('Total Job Growth: ',TenYearTotalGrowth)
TopFiveTotalChange = IndustryAndTotalChange[0][1] + IndustryAndTotalChange[1][1] + IndustryAndTotalChange[2][1] + IndustryAndTotalChange[3][1] + IndustryAndTotalChange[4][1]
print('Top Five Industries Total Job Growth: ',TopFiveTotalChange)
TenYearMinusTopFive = TenYearTotalGrowth - TopFiveTotalChange
print('Remainder: ',TenYearMinusTopFive)

# Data to plot
labels = [IndustryAndTotalChange[0][0], IndustryAndTotalChange[1][0], IndustryAndTotalChange[2][0], IndustryAndTotalChange[3][0], IndustryAndTotalChange[4][0], 'All Others']
sizes = [IndustryAndTotalChange[0][1], IndustryAndTotalChange[1][1], IndustryAndTotalChange[2][1], IndustryAndTotalChange[3][1], IndustryAndTotalChange[4][1], TenYearMinusTopFive]
explode = (0, 0.1, 0, 0, 0.1, 0)  # explode 1st slice

# Plot
plt.pie(sizes, explode=explode, labels=labels, autopct='%1.1f%%', shadow=True, startangle=140)

plt.title('Percentage of 10 Year Job Growth by Industry')
plt.axis('equal')
plt.show()

with open('/Users/bill/School/MSIT/Data/ProjectionsCounts.csv') as MyFile:
    reader = csv.DictReader(MyFile) # comma is default delimiter
    to_db = [(i['agegrpcode'], i['agegrp'], i['yr_2015'], i['yr_2025']) for i in reader]

cur.executemany("INSERT INTO popproj (agegrpcode, agegrp, yr_2015, yr_2025) VALUES (?, ?, ?, ?);", to_db)
con.commit()

wb = openpyxl.load_workbook('/Users/bill/School/MSIT/Data/jobprojectionsoutput.xlsx')

sheet1 = wb.get_sheet_by_name('Sheet1')
wb.remove_sheet(sheet1)
wb.create_sheet('Sheet1')
sheet1 = wb.get_sheet_by_name('Sheet1')

sheet1.cell(row=1, column=1).value = 'Occupation'
sheet1.cell(row=1, column=2).value = '10 Year Projected Job Growth in NYS'

for row in IndustryAndTotalChange:
    sheet1.append(row)

sheet2 = wb.get_sheet_by_name('Sheet2')
wb.remove_sheet(sheet2)
wb.create_sheet('Sheet2')
sheet2 = wb.get_sheet_by_name('Sheet2')

sheet2.cell(row=1, column=1).value = 'Age Group'
sheet2.cell(row=1, column=2).value = '2015'
sheet2.cell(row=1, column=3).value = '2025 Projected'
sheet2.cell(row=1, column=4).value = 'Change in Population'

for row in cur.execute('SELECT agegrp, SUM(yr_2015), SUM(yr_2025), SUM(yr_2025) - SUM(yr_2015) as total FROM popproj GROUP BY agegrpcode ORDER BY agegrpcode'):
    sheet2.append(row)

wb.save('/Users/bill/School/MSIT/Data/jobprojectionsoutput.xlsx')

con.close()
ProjectionsFile.close()
OccupationCodesFile.close()

